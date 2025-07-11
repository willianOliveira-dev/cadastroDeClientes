import customtkinter as ctk
import tkcalendar as tkcal
from time import strftime
from openpyxl import Workbook, load_workbook
from tkinter import messagebox

class App(ctk.CTk):
    
    # Classe principal da aplicação, herda de ctk.CTk
    def __init__(self, fg_color=None, **kwargs):
        super().__init__(fg_color, **kwargs)
        # Configurações iniciais da interface
        self.layoutConfiguracao()  
        # Configuração de aparência
        self.alterarAparencia()
        # Configuração do sistema (frames, labels, tabs, etc.)
        self.sistema()
        # Configuração da hora do sistema
        self.horaSistema()
        

    def layoutConfiguracao(self):
        # Define o tema padrão para o aplicativo (azul)
        ctk.set_default_color_theme('blue')
        # Define o título da janela
        self.title(string='Cadastramento de Clientes')
        # Define o tamanho inicial da janela
        self.geometry(geometry_string='900x620')
        # Probibir redimensionamento
        self.resizable(False, False)
        # Configura a expansão de colunas para melhor ajuste no layout
        for i in range(5):
            self.grid_columnconfigure(i, weight=2)

    # Método para alterar cores com base no modo (claro ou escuro)
    def alterarAparencia(self):
        
        # Função anônima para definir a cor de fundo com base no modo
        colorMode = lambda mode: '#EBEBEB' if mode == 'Light' else '#242424'

        # Verifica widgets específicos (por tipo) dentro da interface
        def checkWidget(widgetCTk: object) -> list:
            return [widget for widget in self.winfo_children() if isinstance(widget, widgetCTk)]

        # Lógica para alterar o modo de aparência
        def change():
            # Verifica o estado do botão switch e define o modo
            mode = 'Light' if switchChangeColor.get() else 'Dark'
            self._set_appearance_mode(mode)  # Define o modo na interface
            
            # Define cores baseadas no modo atual
            newColors = '#000' if switchChangeColor.get() else '#fff'

            # Configura o botão switch
            switchChangeColor.configure(
                text=mode,
                font=('Roboto', 12),
                text_color=newColors,
                bg_color=colorMode(mode=mode),
                
            )
            
            # Altera a cor de todos os labels na interface
            for widgetLabel in checkWidget(ctk.CTkLabel):
                widgetLabel.configure(
                    text_color=newColors, 
                    bg_color=colorMode(mode=mode)
                )
            
            # Altera a cor do primeiro frame encontrado
            frameAdjust = checkWidget(ctk.CTkFrame)
            frameAdjust[0].configure(bg_color=colorMode(mode=mode))
            
            # Configura o TabView e seus widgets
            tabViewAdjust = checkWidget(ctk.CTkTabview)
            if tabViewAdjust:
                tabViewAdjust[0].configure(
                    fg_color='#fff' if mode == 'Light' else '#2b2e2c',
                    bg_color=colorMode(mode=mode)
                )
                
                # Itera por todos os labels dentro das abas do TabView
                for tabviewConfigLabel in tabViewAdjust:
                    for infoLabel in tabviewConfigLabel.winfo_children(): # infoLabel -> Aba1 e Aba2
                        for label in infoLabel.winfo_children(): # Dentro da Aba1 e Aba2,  verifico os labels
                            if isinstance(label, ctk.CTkLabel): # Se pertencer ao objeto CTkLabel ->  configura a cor de texto
                                label.configure(text_color=newColors)
                                
    
        # Cria o botão switch para alternar entre claro e escuro
        switchChangeColor = ctk.CTkSwitch(
            self, 
            text='Tema', 
            font=('Roboto', 12),  
            command=change,
            button_hover_color='#c4003b'
        )
        # Posiciona o botão switch
        switchChangeColor.grid(row=5, column=0, sticky='w', padx=5, pady=5)
        
    def horaSistema(self):
        
        def updateTime():
            currentTime = strftime('%H:%M:%S')
            timeLabel.configure(text=currentTime)
            timeLabel.after(ms=1000, func=updateTime) # Altera o horário a cada 1 segundo
            
        timeLabel = ctk.CTkLabel( #Inicia com um horário
            self, 
            text=strftime('%H:%M:%S'), 
            font=('Coolvetica', 28),
            )
        
        timeLabel.grid(row=2, column=0, padx=15, pady=5, sticky='n')
        updateTime()
 
    # Método que organiza os widgets do sistema
    def sistema(self):
        
        # Função criar arquivo
        def createFile(excelPath):
                try:
                    excelFile = load_workbook(excelPath) 
                except FileNotFoundError:
                    excelFile = Workbook()
                    physicalSheet = excelFile.active
                    physicalSheet.title = 'PessoaFisica'
                    legalSheet = excelFile.create_sheet('PessoaJuridica')
                    
                    physicalSheet['A1'] = nameLabel.cget(attribute_name='text')
                    physicalSheet['B1'] = lastNameLabel.cget(attribute_name='text')
                    physicalSheet['C1'] = genderLabel.cget(attribute_name='text')
                    physicalSheet['D1'] = emailLabel.cget(attribute_name='text')
                    physicalSheet['E1'] = passwordLabel.cget(attribute_name='text')
                    physicalSheet['F1'] = 'Data(Abertura)'
                    physicalSheet['G1'] = cpfLabel.cget(attribute_name='text')
                    physicalSheet['H1'] = phoneLabel.cget(attribute_name='text')
                    physicalSheet['I1'] = addressLabel.cget(attribute_name='text')
                    
                    legalSheet['A1'] = companyNameLabel.cget(attribute_name='text')
                    legalSheet['B1'] = tradeNameLabel.cget(attribute_name='text')
                    legalSheet['C1'] = companyTypeLabel.cget(attribute_name='text')
                    legalSheet['D1'] = companyEmailLabel.cget(attribute_name='text')
                    legalSheet['E1'] = companyPasswordLabel.cget(attribute_name='text')
                    legalSheet['F1'] = 'Data(Abertura)'
                    legalSheet['G1'] = cnpjLabel.cget(attribute_name='text')
                    legalSheet['H1'] = companyPhoneLabel.cget(attribute_name='text')
                    legalSheet['I1'] = companyAddressLabel.cget(attribute_name='text')
                    
                    excelFile.save(excelPath)
                
                return excelFile
            
        # Frame principal do sistema
        frame = ctk.CTkFrame(
            self, 
            width=550,
            height=100,
            fg_color='#c4003b',  # Cor de fundo do frame
        )
        
        # Label com título do sistema
        systemLabel = ctk.CTkLabel(
            frame, 
            text='Sistema de Cadastro de Clientes', 
            text_color='#fff',  # Cor do texto
            font=('Coolvetica', 36),  # Fonte e tamanho
            bg_color='transparent',  # Cor de fundo transparente
            height=60
        )
        
        # Configuração da expansão da coluna para centralizar
        frame.grid_columnconfigure(0, weight=1)
        # Posiciona o título no frame
        systemLabel.grid(row=0, column=0, padx=10, pady=10)
        # Posiciona o frame no grid principal
        frame.grid(row=0, column=0, columnspan=6, padx=10, pady=10, sticky='nsew')

        # Label "Cadastra-se" abaixo do frame principal
        registerLabel = ctk.CTkLabel(
            self, 
            text='Cadastra-se',
            text_color='#fff',  # Cor do texto
            font=('Coolvetica', 32),  # Fonte e tamanho
            bg_color='transparent',  # Cor de fundo
            anchor='w'  # Alinhamento à esquerda
        )
        # Posiciona o label no grid
        registerLabel.grid(row=1, column=2, padx=10, pady=10, sticky='nw')
        
        # Cria o TabView para as abas de cadastro
        tabsView = ctk.CTkTabview(
            self, 
            width=800, 
            height=400, 
            border_width=2,  # Largura da borda
            segmented_button_selected_color='#c4003b',  # Cor da aba selecionada
            segmented_button_selected_hover_color='#730732',  # Cor ao passar o mouse
            border_color='#c4003b',  # Cor da borda
            segmented_button_fg_color='#000'
        )
        # Posiciona o TabView no grid
        tabsView.grid(row=2, column=0, columnspan=5, padx=5, pady=5)

        # Adiciona abas no TabView
        tabsView.add('Pessoa Física')
        tabsView.add('Pessoa Jurídica')
        
        # Adiciona labels dentro da aba "Pessoa Física"
        
        # Linha 0 e 1
        # Coluna 0 e 3
        # Nome
        nameLabel = ctk.CTkLabel(tabsView.tab('Pessoa Física'), text='Nome', font=('Roboto', 16))
        nameLabel.grid(row=0, column=0, padx=15, pady=5, sticky='nw')
        
        nameEntry = ctk.CTkEntry(tabsView.tab('Pessoa Física'), placeholder_text='Digite o seu Nome', width=250)
        nameEntry.grid(row=1, column=0, columnspan=2, padx=10, pady=5, sticky='nw')
        
        # Sobrenome
        lastNameLabel = ctk.CTkLabel(tabsView.tab('Pessoa Física'), text='Sobrenome', font=('Roboto', 16))
        lastNameLabel.grid(row=0, column=3, padx=15, pady=5, sticky='nw')
        
        lastNameEntry = ctk.CTkEntry(tabsView.tab('Pessoa Física'), placeholder_text='Digite o seu Sobrenome', width=250)
        lastNameEntry.grid(row=1, column=3, columnspan=2, padx=10, pady=5, sticky='nw')
        
        # Opção de Menu
        genderLabel = ctk.CTkLabel(tabsView.tab('Pessoa Física'), text='Sexo', font=('Roboto', 16))
        genderLabel.grid(row=0, column=5, padx=15, pady=5, sticky='nw')
        
        genderMenu = ctk.CTkOptionMenu(
            tabsView.tab('Pessoa Física'), 
            values=['Masculino', 'Feminino', 'Outros'], 
            bg_color='transparent', 
            fg_color='#c4003b', 
            dropdown_hover_color='#c4003b'
            )
        
        genderMenu.grid(row=1, column=5, padx=10, pady=5, sticky='nw')
        
        # Linha 2 e 3
        # Coluna 0 e 3
        # Email
        emailLabel = ctk.CTkLabel(tabsView.tab('Pessoa Física'), text='E-mail', font=('Roboto', 16))
        emailLabel.grid(row=2, column=0, padx=15, pady=5, sticky='nw')
        
        emailEntry = ctk.CTkEntry(tabsView.tab('Pessoa Física'), placeholder_text='Digite o seu E-mail', width=250)
        emailEntry.grid(row=3, column=0, columnspan=2, padx=10, pady=5, sticky='nw')
        
        # Senha
        passwordLabel = ctk.CTkLabel(tabsView.tab('Pessoa Física'), text='Senha', font=('Roboto', 16))
        passwordLabel.grid(row=2, column=3, padx=15, pady=5, sticky='nw')
        
        passwordEntry = ctk.CTkEntry(tabsView.tab('Pessoa Física'), placeholder_text='Digite a sua Senha', show='*', width=250)
        passwordEntry.grid(row=3, column=3, columnspan=2, padx=10, pady=5, sticky='nw')
        
        #data
        openingDateLabel = ctk.CTkLabel(tabsView.tab('Pessoa Física'), text='Data(Abertura)', font=('Roboto', 16))
        openingDateLabel.grid(row=2, column=5, padx=15, pady=5, sticky='nw')
        
        calendar = tkcal.DateEntry(tabsView.tab('Pessoa Física'), font=('Roboto', 12), width=15)
        calendar.grid(row=3, column=5, padx=15, pady=5, sticky='nw')
    
        # Linha 4 e 5
        # Coluna 0 e 3
        # CPF
        cpfLabel = ctk.CTkLabel(tabsView.tab('Pessoa Física'), text='CPF', font=('Roboto', 16))
        cpfLabel.grid(row=4, column=0, padx=15, pady=5, sticky='nw')
        
        cpfEntry = ctk.CTkEntry(tabsView.tab('Pessoa Física'), placeholder_text='Informe apenas dígitos', width=250)
        cpfEntry.grid(row=5, column=0, columnspan=2, padx=10, pady=5, sticky='nw')
        
        # telefone
        phoneLabel = ctk.CTkLabel(tabsView.tab('Pessoa Física'), text='Telefone', font=('Roboto', 16))
        phoneLabel.grid(row=4, column=3, padx=15, pady=5, sticky='nw')
        
        phoneEntry = ctk.CTkEntry(tabsView.tab('Pessoa Física'), placeholder_text='Telefone', width=250)
        phoneEntry.grid(row=5, column=3, columnspan=2, padx=10, pady=5, sticky='nw')
        
        # Endereço
        addressLabel = ctk.CTkLabel(tabsView.tab('Pessoa Física'), text='Endereço', font=('Roboto', 16))
        addressLabel.grid(row=4, column=5, padx=15, pady=5, sticky='nw')
        
        addressEntry = ctk.CTkEntry(tabsView.tab('Pessoa Física'), placeholder_text='Digite o seu Endereço', width=250)
        addressEntry.grid(row=5, column=5, columnspan=2, padx=10, pady=5, sticky='nw')
        
        # Pessoa física
        # Botões de Ação - Importante para executar as principais funcionalidades
        
        # Função limpar
        def clearPhysicalFields():
            for widgetTabview in self.winfo_children():
                if isinstance(widgetTabview, ctk.CTkTabview):
                    physicalTab = widgetTabview.winfo_children()[0]
                    for entry in physicalTab.winfo_children():
                        if isinstance(entry, ctk.CTkEntry):
                            entry.delete(0, ctk.END)
                            passwordEntry.configure(show='*') # Segurança -  Manter a proteção das informações inseridas no campo senha
                     
        # Botão Limpar campos               
        clearPhysicalButton = ctk.CTkButton(
            tabsView.tab('Pessoa Física'), 
            text='Limpar Campos', 
            font=('Roboto', 16),
            height=40,
            width=250,
            text_color='#fff',
            bg_color='transparent',
            fg_color='#c4003b',
            hover_color='#730732',
            command=clearPhysicalFields
            )
        
        clearPhysicalButton.grid(row=6, column=0, padx=15, pady=15, sticky='sw')
         
         # Função enviar dados
        def sendPhysicalData():
            
            def registerClients(excelFilePath, *args):
                excelFile = load_workbook(excelFilePath)
                worksheet = excelFile['PessoaFisica'] # ativa nossa planilha
                newRow = worksheet.max_row # pega número da última linha
                
                columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
                for i, info in enumerate(args):
                    if i < len(columns):  # Verifica se o número de argumentos não excede as colunas
                        worksheet[f'{columns[i]}{newRow + 1}'] = info
                    
                excelFile.save(excelFilePath)
         
            name = nameEntry.get()
            lastName = lastNameEntry.get()
            gender = genderMenu.get()
            email = emailEntry.get()
            password = passwordEntry.get()
            date = calendar.get_date()
            cpf = cpfEntry.get()
            phone = phoneEntry.get()
            address = addressEntry.get()

            # Chama a função para criar ou abrir o arquivo
            filePath = 'cadastroClientes.xlsx'
            createFile(filePath) # Validação que o nosso arquivo existe
            
            registerClients(filePath, name, lastName, gender, email, password, date, cpf, phone, address)
            
            messagebox.showinfo(title='Mensagem sobre Cadastro', message='Cliente cadastrado com sucesso!')          
            
        # Botão Enviar
        sendPhysicalButton = ctk.CTkButton(
            tabsView.tab('Pessoa Física'), 
            text='Enviar', 
            font=('Roboto', 16),
            height=40,
            width=250,
            text_color='#fff',
            bg_color='transparent',
            fg_color='#02ad10',
            hover_color='#730732',
            command=sendPhysicalData
            )
        
        sendPhysicalButton.grid(row=6, column=3, padx=15, pady=15, sticky='sw')
    
        # Pessoa Jurídica TODO
        
         # Linha 0 e 1
        # Coluna 0 e 3
        # Nome Empresa
        companyNameLabel = ctk.CTkLabel(tabsView.tab('Pessoa Jurídica'), text='Empresa', font=('Roboto', 16))
        companyNameLabel.grid(row=0, column=0, padx=15, pady=5, sticky='nw')
        
        companyNameEntry = ctk.CTkEntry(tabsView.tab('Pessoa Jurídica'), placeholder_text='Nome Empresarial', width=250)
        companyNameEntry.grid(row=1, column=0, columnspan=2, padx=10, pady=5, sticky='nw')
        
        # TÍTULO DO ESTABELECIMENTO ( NOME FANTASIA)
        tradeNameLabel = ctk.CTkLabel(tabsView.tab('Pessoa Jurídica'), text='Nome Fantasia', font=('Roboto', 16))
        tradeNameLabel.grid(row=0, column=3, padx=15, pady=5, sticky='nw')
        
        tradeNameEntry = ctk.CTkEntry(tabsView.tab('Pessoa Jurídica'), placeholder_text='Informe o Título do Estabelecimento', width=250)
        tradeNameEntry.grid(row=1, column=3, columnspan=2, padx=10, pady=5, sticky='nw')
        
        # Opção de Menu
        companyTypeLabel = ctk.CTkLabel(tabsView.tab('Pessoa Jurídica'), text='Tipo de Empresa', font=('Roboto', 16))
        companyTypeLabel.grid(row=0, column=5, padx=15, pady=5, sticky='nw')
        
        companyTypeMenu = ctk.CTkOptionMenu(
            tabsView.tab('Pessoa Jurídica'), 
            values=['Microempreendedor Individual (MEI)', 'Microempresa (ME)', 'Sociedade Anônima (S/A)'], 
            bg_color='transparent', 
            fg_color='#c4003b', 
            dropdown_hover_color='#c4003b'
            )
        
        companyTypeMenu.grid(row=1, column=5, padx=10, pady=5, sticky='nw')
        
        # Linha 2 e 3
        # Coluna 0 e 3
        # Email Empresarial
        companyEmailLabel = ctk.CTkLabel(tabsView.tab('Pessoa Jurídica'), text='E-mail Empresarial', font=('Roboto', 16))
        companyEmailLabel.grid(row=2, column=0, padx=15, pady=5, sticky='nw')
        
        companyEmailEntry = ctk.CTkEntry(tabsView.tab('Pessoa Jurídica'), placeholder_text='Digite o E-mail da Empresa ', width=250)
        companyEmailEntry.grid(row=3, column=0, columnspan=2, padx=10, pady=5, sticky='nw')
        
        # Senha
        companyPasswordLabel = ctk.CTkLabel(tabsView.tab('Pessoa Jurídica'), text='Senha', font=('Roboto', 16))
        companyPasswordLabel.grid(row=2, column=3, padx=15, pady=5, sticky='nw')
        
        companyPasswordEntry = ctk.CTkEntry(tabsView.tab('Pessoa Jurídica'), placeholder_text='Digite a sua Senha', show='*', width=250)
        companyPasswordEntry.grid(row=3, column=3, columnspan=2, padx=10, pady=5, sticky='nw')
        
        #data
        companyOpeningDateLabel = ctk.CTkLabel(tabsView.tab('Pessoa Jurídica'), text='Data(Abertura)', font=('Roboto', 16))
        companyOpeningDateLabel.grid(row=2, column=5, padx=15, pady=5, sticky='nw')
        
        # Calendario 
        companyCalendar = tkcal.DateEntry(tabsView.tab('Pessoa Jurídica'), font=('Roboto', 12), width=15)
        companyCalendar.grid(row=3, column=5, padx=15, pady=5, sticky='nw')
    
        # Linha 4 e 5
        # Coluna 0 e 3
        # CNPJ
        cnpjLabel = ctk.CTkLabel(tabsView.tab('Pessoa Jurídica'), text='CNPJ', font=('Roboto', 16))
        cnpjLabel.grid(row=4, column=0, padx=15, pady=5, sticky='nw')
        
        cnpjEntry = ctk.CTkEntry(tabsView.tab('Pessoa Jurídica'), placeholder_text='Informe apenas dígitos', width=250)
        cnpjEntry.grid(row=5, column=0, columnspan=2, padx=10, pady=5, sticky='nw')
        
        # telefone
        companyPhoneLabel = ctk.CTkLabel(tabsView.tab('Pessoa Jurídica'), text='Telefone', font=('Roboto', 16))
        companyPhoneLabel.grid(row=4, column=3, padx=15, pady=5, sticky='nw')
        
        companyPhoneEntry = ctk.CTkEntry(tabsView.tab('Pessoa Jurídica'), placeholder_text='Telefone', width=250)
        companyPhoneEntry.grid(row=5, column=3, columnspan=2, padx=10, pady=5, sticky='nw')
        
        # Endereço
        companyAddressLabel = ctk.CTkLabel(tabsView.tab('Pessoa Jurídica'), text='Endereço', font=('Roboto', 16))
        companyAddressLabel.grid(row=4, column=5, padx=15, pady=5, sticky='nw')
        
        companyAddressEntry = ctk.CTkEntry(tabsView.tab('Pessoa Jurídica'), placeholder_text='Digite o Endereço do Estabelecimento', width=250)
        companyAddressEntry.grid(row=5, column=5, columnspan=2, padx=10, pady=5, sticky='nw')
        
        # Botões de Ação - Importante para executar as principais funcionalidades
        
        # Função Limpar
        def clearLegalFields():
            for widgetTabview in self.winfo_children():
                if isinstance(widgetTabview, ctk.CTkTabview):
                    legalTab = widgetTabview.winfo_children()[1]
                    for entry in legalTab.winfo_children():
                        if isinstance(entry, ctk.CTkEntry):
                            entry.delete(0, ctk.END)
                            companyPasswordEntry.configure(show='*') # Segurança -  Manter a proteção das informações inseridas no campo senha
                                    
        # Botão limmpar               
        clearLegalButton = ctk.CTkButton(
            tabsView.tab('Pessoa Jurídica'), 
            text='Limpar Campos', 
            font=('Roboto', 16),
            height=40,
            width=250,
            text_color='#fff',
            bg_color='transparent',
            fg_color='#c4003b',
            hover_color='#730732',
            command=clearLegalFields
            )
        
        clearLegalButton.grid(row=6, column=0, padx=15, pady=15, sticky='sw')
    
        # Função Enviar dados
        def sendLegalData():
        
            def registerClients(excelFilePath, *args):
                excelFile = load_workbook(excelFilePath)
                worksheet = excelFile['PessoaJuridica'] # ativa nossa planilha
                newRow = worksheet.max_row # pega número da última linha
                
                columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
                for i, info in enumerate(args):
                    if i < len(columns):  # Verifica se o número de argumentos não excede as colunas
                        worksheet[f'{columns[i]}{newRow + 1}'] = info
                    
                excelFile.save(excelFilePath)
         
            companyName = companyNameEntry.get()
            tradeName = tradeNameEntry.get()
            companyType = companyTypeMenu.get()
            companyEmail = companyEmailEntry.get()
            companyPassword = companyPasswordEntry.get()
            companyDate = companyCalendar.get_date()
            cnpj = cnpjEntry.get()
            companyPhone = companyPhoneEntry.get()
            companyAddress = companyAddressEntry.get()

            # Chama a função para criar ou abrir o arquivo
            filePath = 'cadastroClientes.xlsx'
            createFile(filePath) # Validação que o nosso arquivo existe
            
            registerClients(filePath, companyName, tradeName, companyType, companyEmail, companyPassword, companyDate, cnpj, companyPhone, companyAddress)
            
            messagebox.showinfo(title='Mensagem sobre Cadastro', message='Cliente cadastrado com sucesso!')         
        sendLegalButton = ctk.CTkButton(
            tabsView.tab('Pessoa Jurídica'), 
            text='Enviar', 
            font=('Roboto', 16),
            height=40,
            width=250,
            text_color='#fff',
            bg_color='transparent',
            fg_color='#02ad10',
            hover_color='#730732',
            command=sendLegalData
            )
        
        sendLegalButton.grid(row=6, column=3, padx=15, pady=15, sticky='sw')
        
        
        # Botão de fechar
        def close():
            self.quit()
        
        for register in ['Pessoa Física', 'Pessoa Jurídica']:
            closeButton = ctk.CTkButton(
                tabsView.tab(register), 
                text='Fechar', 
                font=('Roboto', 16),
                height=30,
                width=100,
                text_color='#fff',
                bg_color='transparent',
                fg_color='#c4003b',
                hover_color='#730732',
                command=close
                )
            
            closeButton.grid(row=7, column=5, padx=10, pady=15, sticky='se')

# Inicializa o aplicativo
if __name__ == '__main__':
    app = App()
    app.mainloop()
